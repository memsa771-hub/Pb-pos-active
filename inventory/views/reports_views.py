from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.db.models import Sum, F, DecimalField
from django.db.models.functions import TruncDay
from django.http import HttpResponse, JsonResponse
from django.utils import timezone
from datetime import timedelta
import json
import io

# Import Models and Decorators
from inventory.models import InventoryItem, Purchase, StockAdjustment
from posapp.decorators import can_view_stock_required

# Excel Export Setup (Optional Dependencies)
EXCEL_EXPORT_AVAILABLE = False
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter
    EXCEL_EXPORT_AVAILABLE = True
except ImportError:
    pass

# ----------------- MAIN REPORTS DASHBOARD ----------------- #

@login_required
@can_view_stock_required
def reports_dashboard(request):
    """
    Shows High-Level Inventory Stats:
    - Total Asset Value (Current Stock * Cost)
    - Monthly Spending
    - Monthly Loss (Wastage)
    """
    
    # 1. Total Inventory Valuation (Assets)
    # Formula: Sum(Cost Price * Current Stock)
    stock_valuation = InventoryItem.objects.filter(is_active=True).aggregate(
        total_value=Sum(F('cost_price') * F('current_stock'), output_field=DecimalField())
    )['total_value'] or 0

    # 2. Monthly Purchase Stats (Current Month)
    today = timezone.now()
    first_day_month = today.replace(day=1, hour=0, minute=0, second=0)
    
    monthly_purchases = Purchase.objects.filter(
        purchase_date__gte=first_day_month
    ).aggregate(total=Sum('total_amount'))['total'] or 0

    # 3. Monthly Wastage Stats
    # Hum sirf 'Wastage' aur 'Theft' ko loss count krain gay
    monthly_wastage_items = StockAdjustment.objects.filter(
        date__gte=first_day_month,
        adjustment_type__in=['Wastage', 'Theft'],
        action='REMOVE'
    )
    
    # Calculate wastage value manually (Qty * Item Cost)
    monthly_wastage_value = sum(adj.quantity * adj.inventory_item.cost_price for adj in monthly_wastage_items)

    # 4. Low Stock Alerts Count
    low_stock_count = 0
    active_items = InventoryItem.objects.filter(is_active=True)
    for item in active_items:
        if item.current_stock <= item.min_stock_level:
            low_stock_count += 1

    context = {
        'stock_valuation': stock_valuation,
        'monthly_purchases': monthly_purchases,
        'monthly_wastage_value': monthly_wastage_value,
        'low_stock_count': low_stock_count,
        'excel_export_available': EXCEL_EXPORT_AVAILABLE
    }
    return render(request, 'inventory/reports/dashboard.html', context)


# ----------------- PURCHASE HISTORY (CHARTS) ----------------- #

@login_required
@can_view_stock_required
def purchase_report(request):
    """Shows Purchase History filtered by date with Chart Data"""
    
    # Date Filtering
    date_range = request.GET.get('range', '30days')
    today = timezone.now().date()
    
    if date_range == '7days':
        start_date = today - timedelta(days=7)
    elif date_range == '30days':
        start_date = today - timedelta(days=30)
    elif date_range == 'this_year':
        start_date = today.replace(month=1, day=1)
    else:
        start_date = today - timedelta(days=30)

    # Query Purchases
    purchases = Purchase.objects.filter(purchase_date__gte=start_date).order_by('purchase_date')

    # Chart Data Preparation (Group by Day)
    daily_data = purchases.values('purchase_date').annotate(
        total=Sum('total_amount')
    ).order_by('purchase_date')

    # Chart Labels
    chart_labels = [item['purchase_date'].strftime('%Y-%m-%d') for item in daily_data]
    chart_values = [float(item['total']) for item in daily_data]

    context = {
        'purchases': purchases,
        'chart_labels': json.dumps(chart_labels),
        'chart_values': json.dumps(chart_values),
        'date_range': date_range,
        'total_spent': sum(chart_values)
    }
    return render(request, 'inventory/reports/purchase_report.html', context)


# ----------------- EXCEL EXPORT (STOCK VALUATION) ----------------- #

@login_required
@can_view_stock_required
def export_inventory_excel(request):
    """Generates an Excel file of Current Inventory Status & Value"""
    
    if not EXCEL_EXPORT_AVAILABLE:
        return JsonResponse({'error': 'Excel dependencies (openpyxl) not installed on server.'}, status=400)

    try:
        # Fetch Active Items
        items = InventoryItem.objects.filter(is_active=True).select_related('category').order_by('category__name', 'name')
        
        # Setup Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Inventory Valuation"
        
        # Styling
        bold = Font(bold=True)
        header_fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
        center = Alignment(horizontal="center", vertical="center")
        
        # Report Header
        ws.merge_cells('A1:E1')
        ws['A1'] = f"Inventory Valuation Report - {timezone.now().strftime('%d %b %Y')}"
        ws['A1'].font = Font(size=14, bold=True)
        ws['A1'].alignment = center
        
        # Table Columns
        headers = ['Category', 'Item Name', 'Current Stock', 'Cost Price (Per Unit)', 'Total Asset Value']
        ws.append([]) # Blank row for spacing
        ws.append(headers)
        
        # Style Headers
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col_num)
            cell.font = bold
            cell.fill = header_fill
            ws.column_dimensions[get_column_letter(col_num)].width = 22

        # Fill Data Rows
        grand_total_value = 0
        
        for item in items:
            total_val = item.current_stock * item.cost_price
            grand_total_value += total_val
            
            ws.append([
                item.category.name if item.category else "Uncategorized",
                item.name,
                f"{item.current_stock} {item.unit}",
                float(item.cost_price),
                float(total_val)
            ])
            
        # Grand Total Row at Bottom
        total_row = ws.max_row + 2
        ws.cell(row=total_row, column=4, value="GRAND TOTAL ASSETS:").font = bold
        ws.cell(row=total_row, column=5, value=float(grand_total_value)).font = bold
        
        # Save to Memory buffer
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Create HTTP Response
        filename = f"Inventory_Valuation_{timezone.now().strftime('%Y%m%d')}.xlsx"
        response = HttpResponse(output.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        return response

    except Exception as e:
        return JsonResponse({'error': f'Export failed: {str(e)}'}, status=500)