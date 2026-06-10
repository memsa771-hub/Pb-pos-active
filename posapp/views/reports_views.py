from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.db.models import Sum, Count, F, DecimalField, Value
from django.db.models.functions import TruncDay, TruncWeek, TruncMonth, Coalesce
from django.http import HttpResponse, JsonResponse
from django.utils import timezone
from django.contrib import messages
from datetime import datetime, timedelta
from decimal import Decimal
from io import BytesIO
import csv
import json
from django.db.models import Q
from django.core.cache import cache
import logging
from ..decorators import management_required, admin_required
from ..permissions import is_admin, is_branch_manager, can_access_management
from ..models import Order, OrderItem, PosProduct, PosCategory, BusinessSettings, BillAdjustment, AdvanceAdjustment, BusinessLogo, Setting, EndDay

import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

# Try to import xlwt for Excel export, but make it optional
EXCEL_EXPORT_AVAILABLE = False
try:
    import pandas as pd
    import openpyxl
    EXCEL_EXPORT_AVAILABLE = True
except ImportError:
    print("Excel export dependencies (pandas, openpyxl) not installed. Excel export will be disabled.")
except Exception as e:
    print(f"Excel export setup error: {e}")

# Set up logger
logger = logging.getLogger('posapp')

def get_business_day_ranges():
    """
    Calculate business day ranges based on End Day cycles instead of calendar days.
    Returns date ranges for 'today', 'last_7_business_days', and 'last_30_business_days'.
    """
    now = timezone.now()
    
    # Get all end days ordered by date
    end_days = list(EndDay.objects.all().order_by('end_date'))
    
    if not end_days:
        # If no end days exist, fall back to calendar-based calculation
        today_start = now.replace(hour=0, minute=0, second=0, microsecond=0)
        return {
            'today': (today_start, now),
            'last_7_business_days': (now - timedelta(days=7), now),
            'last_30_business_days': (now - timedelta(days=30), now),
            'current_business_period': (now - timedelta(days=30), now)
        }
    
    # Current business day: from last end day to now
    last_end_day = end_days[-1]
    current_business_day_start = last_end_day.end_date
    
    # Find the 7th and 30th business days back from current
    # We count backwards through end day cycles
    
    # For "last 7 business days": get the start of the 7th business day back
    if len(end_days) >= 7:
        seventh_day_back = end_days[-7]
        # Start from the end day before the 7th day back
        if len(end_days) >= 8:
            last_7_start = end_days[-8].end_date
        else:
            # If we don't have 8 end days, use the earliest order or 30 days ago
            earliest_order = Order.objects.order_by('created_at').first()
            if earliest_order:
                last_7_start = earliest_order.created_at
            else:
                last_7_start = now - timedelta(days=30)
    else:
        # If we have fewer than 7 business days, use all available data
        if len(end_days) >= 2:
            last_7_start = end_days[0].end_date
        else:
            # Use earliest order or 30 days ago
            earliest_order = Order.objects.order_by('created_at').first()
            if earliest_order:
                last_7_start = earliest_order.created_at
            else:
                last_7_start = now - timedelta(days=30)
    
    # For "last 30 business days": get the start of the 30th business day back
    if len(end_days) >= 30:
        thirtieth_day_back = end_days[-30]
        # Start from the end day before the 30th day back
        if len(end_days) >= 31:
            last_30_start = end_days[-31].end_date
        else:
            # If we don't have 31 end days, use the earliest order
            earliest_order = Order.objects.order_by('created_at').first()
            if earliest_order:
                last_30_start = earliest_order.created_at
            else:
                last_30_start = now - timedelta(days=60)
    else:
        # If we have fewer than 30 business days, use all available data
        if len(end_days) >= 2:
            last_30_start = end_days[0].end_date
        else:
            # Use earliest order or 60 days ago
            earliest_order = Order.objects.order_by('created_at').first()
            if earliest_order:
                last_30_start = earliest_order.created_at
            else:
                last_30_start = now - timedelta(days=60)
    
    return {
        'today': (current_business_day_start, now),
        'last_7_business_days': (last_7_start, now),
        'last_30_business_days': (last_30_start, now),
        'current_business_period': (current_business_day_start, now)
    }

@login_required
@management_required
def reports_dashboard(request):
    """Main reports dashboard with overview of available reports"""
    
    # Check if user is admin or branch manager
    user_is_admin = is_admin(request.user)
    
    # Get the last end day timestamp
    last_end_day = EndDay.get_last_end_day()
    last_end_day_time = last_end_day.end_date if last_end_day else None
    
    # Get business day ranges based on End Day cycles
    business_ranges = get_business_day_ranges()
    
    # Base query for filtering orders
    if user_is_admin or not last_end_day_time:
        # Admin sees all orders, or if no end day exists, show all
        base_query = Order.objects.all()
    else:
        # Branch manager sees only orders since last end day
        base_query = Order.objects.filter(created_at__gte=last_end_day_time)
    
    # Orders count (exclude cancelled orders) - using business day ranges
    today_start, today_end = business_ranges['today']
    week_start, week_end = business_ranges['last_7_business_days']
    month_start, month_end = business_ranges['last_30_business_days']
    
    orders_today = base_query.filter(created_at__gte=today_start, created_at__lte=today_end).exclude(order_status='Cancelled').count()
    orders_week = base_query.filter(created_at__gte=week_start, created_at__lte=week_end).exclude(order_status='Cancelled').count()
    orders_month = base_query.filter(created_at__gte=month_start, created_at__lte=month_end).exclude(order_status='Cancelled').count()
    orders_total = base_query.exclude(order_status='Cancelled').count()
    
    # Revenue (exclude cancelled orders) - using business day ranges
    revenue_today = base_query.filter(created_at__gte=today_start, created_at__lte=today_end).exclude(order_status='Cancelled').aggregate(total=Sum('total_amount'))['total'] or 0
    revenue_week = base_query.filter(created_at__gte=week_start, created_at__lte=week_end).exclude(order_status='Cancelled').aggregate(total=Sum('total_amount'))['total'] or 0
    revenue_month = base_query.filter(created_at__gte=month_start, created_at__lte=month_end).exclude(order_status='Cancelled').aggregate(total=Sum('total_amount'))['total'] or 0
    revenue_total = base_query.exclude(order_status='Cancelled').aggregate(total=Sum('total_amount'))['total'] or 0
    
    # Get categories for the product export filter
    categories = PosCategory.objects.all()
    
    context = {
        'orders_today': orders_today,
        'orders_week': orders_week,
        'orders_month': orders_month,
        'orders_total': orders_total,
        'revenue_today': revenue_today,
        'revenue_week': revenue_week,
        'revenue_month': revenue_month,
        'revenue_total': revenue_total,
        'categories': categories,
        'excel_export_available': EXCEL_EXPORT_AVAILABLE,
        'is_admin': user_is_admin,
        'last_end_day': last_end_day,
        # Add business day range info for template display
        'business_ranges': business_ranges,
        'today_range': f"{today_start.strftime('%b %d')} - {today_end.strftime('%b %d')}",
        'week_range': f"{week_start.strftime('%b %d')} - {week_end.strftime('%b %d')}",
        'month_range': f"{month_start.strftime('%b %d')} - {month_end.strftime('%b %d')}",
    }
    
    return render(request, 'posapp/reports/dashboard.html', context)


@login_required
@management_required
def sales_report(request):
    """Sales report with charts and data"""
    
    # Check if user is admin or branch manager
    user_is_admin = is_admin(request.user)
    
    # Get the last end day timestamp
    last_end_day = EndDay.get_last_end_day()
    last_end_day_time = last_end_day.end_date if last_end_day else None
    
    # Handle date range selection
    report_type = request.GET.get('report_type', 'daily')
    date_range = request.GET.get('date_range', '7days')
    custom_start = request.GET.get('start', '')
    custom_end = request.GET.get('end', '')
    
    # Calculate date range based on selection
    today = timezone.now().date()
    
    # For business day based ranges (7days and 30days), use End Day cycles
    if date_range == '7days':
        business_ranges = get_business_day_ranges()
        start_datetime, end_datetime = business_ranges['last_7_business_days']
        start_date = start_datetime.date()
        end_date = end_datetime.date()
    elif date_range == '30days':
        business_ranges = get_business_day_ranges()
        start_datetime, end_datetime = business_ranges['last_30_business_days']
        start_date = start_datetime.date()
        end_date = end_datetime.date()
    elif date_range == 'this_month':
        start_date = today.replace(day=1)
        end_date = today
    elif date_range == 'last_month':
        last_month = today.month - 1 if today.month > 1 else 12
        last_month_year = today.year if today.month > 1 else today.year - 1
        start_date = datetime(last_month_year, last_month, 1).date()
        if last_month == 12:
            end_date = datetime(last_month_year, last_month, 31).date()
        else:
            end_date = datetime(last_month_year, last_month + 1, 1).date() - timedelta(days=1)
    elif date_range == 'this_year':
        start_date = today.replace(month=1, day=1)
        end_date = today
    elif date_range == 'custom' and custom_start and custom_end:
        start_date = datetime.strptime(custom_start, '%Y-%m-%d').date()
        end_date = datetime.strptime(custom_end, '%Y-%m-%d').date()
    else:
        # Default to last 7 business days
        business_ranges = get_business_day_ranges()
        start_datetime, end_datetime = business_ranges['last_7_business_days']
        start_date = start_datetime.date()
        end_date = end_datetime.date()
    
    # Get sales data grouped by day/week/month
    if report_type == 'daily':
        truncate_date = TruncDay('created_at')
    elif report_type == 'weekly':
        truncate_date = TruncWeek('created_at')
    else:  # monthly
        truncate_date = TruncMonth('created_at')
    
    # For branch managers, further limit by last end day if needed
    base_filter = Q(created_at__date__gte=start_date, created_at__date__lte=end_date)
    
    if not user_is_admin and last_end_day_time:
        # Add condition for branch managers to only see data since last end day
        base_filter &= Q(created_at__gte=last_end_day_time)
    
    # Get sales data excluding cancelled orders
    sales_data = Order.objects.filter(
        base_filter
    ).exclude(
        order_status='Cancelled'
    ).annotate(
        date=truncate_date
    ).values('date').annotate(
        total_sales=Sum('total_amount'),
        order_count=Count('id')
    ).order_by('date')
    
    # Top selling products (exclude orders that were cancelled)
    top_products = OrderItem.objects.filter(
        order__in=Order.objects.filter(base_filter),
        order__order_status='Completed'  # Only include completed orders
    ).values('product__name', 'product__category__name').annotate(
        total_quantity=Sum('quantity'),
        total_sales=Sum('total_price')
    ).order_by('-total_quantity')[:10]
    
    # Sales by category (exclude orders that were cancelled)
    category_sales = OrderItem.objects.filter(
        order__in=Order.objects.filter(base_filter)
    ).exclude(
        order__order_status='Cancelled'
    ).values('product__category__name').annotate(
        total_sales=Sum('total_price')
    ).order_by('-total_sales')
    
    # Prepare chart data
    chart_labels = []
    chart_sales = []
    chart_orders = []
    
    for item in sales_data:
        if report_type == 'daily':
            chart_labels.append(item['date'].strftime('%Y-%m-%d'))
        elif report_type == 'weekly':
            chart_labels.append(f"Week {item['date'].strftime('%U')}")
        elif report_type == 'monthly':
            chart_labels.append(item['date'].strftime('%b %Y'))
        
        chart_sales.append(float(item['total_sales']))
        chart_orders.append(item['order_count'])
    
    # Category pie chart data
    category_labels = [item['product__category__name'] or 'Unknown' for item in category_sales]
    category_data = [float(item['total_sales']) for item in category_sales]
    
    # Add business day info for template display
    is_business_day_range = date_range in ['7days', '30days']
    range_description = ""
    if is_business_day_range:
        if date_range == '7days':
            range_description = "Last 7 Business Days (End Day Cycles)"
        elif date_range == '30days':
            range_description = "Last 30 Business Days (End Day Cycles)"
    
    context = {
        'report_type': report_type,
        'date_range': date_range,
        'custom_start': custom_start,
        'custom_end': custom_end,
        'start_date': start_date,
        'end_date': end_date,
        'sales_data': sales_data,
        'top_products': top_products,
        'category_sales': category_sales,
        'chart_labels': json.dumps(chart_labels),
        'chart_sales': json.dumps(chart_sales),
        'chart_orders': json.dumps(chart_orders),
        'category_labels': json.dumps(category_labels),
        'category_data': json.dumps(category_data),
        'total_sales': sum(chart_sales),
        'total_orders': sum(chart_orders),
        'excel_export_available': EXCEL_EXPORT_AVAILABLE,
        'is_admin': user_is_admin,
        'last_end_day': last_end_day,
        'is_business_day_range': is_business_day_range,
        'range_description': range_description,
    }
    
    return render(request, 'posapp/reports/sales_report.html', context)


@login_required
@management_required
def export_orders_excel(request):
    """Export orders as modern Excel (.xlsx) file using pandas + openpyxl"""

    if not EXCEL_EXPORT_AVAILABLE:
        return JsonResponse({
            'error': 'Excel export functionality requires pandas and openpyxl packages.'
        }, status=400)

    # Get filter parameters
    start_date = request.GET.get('start')
    end_date = request.GET.get('end')
    status = request.GET.get('status')

    today = timezone.now()

    try:
        # Parse and validate start/end dates
        def parse_date(value, default, end=False):
            if not value:
                return default
            try:
                dt = datetime.strptime(value, '%Y-%m-%d %H:%M:%S')
            except ValueError:
                dt = datetime.strptime(value, '%Y-%m-%d')
                if not end:
                    dt = datetime.combine(dt, datetime.min.time())
                else:
                    dt = datetime.combine(dt, datetime.max.time())
            if timezone.is_naive(dt):
                dt = timezone.make_aware(dt)
            return dt

        start_date_obj = parse_date(start_date, today - timedelta(days=30))
        end_date_obj = parse_date(end_date, today, end=True)

        if start_date_obj > end_date_obj:
            return JsonResponse({'error': 'Start date cannot be after end date.'}, status=400)

        # Filter orders
        orders = Order.objects.filter(created_at__gte=start_date_obj, created_at__lte=end_date_obj)
        if status:
            orders = orders.filter(order_status=status)

        if not orders.exists():
            return JsonResponse({'error': 'No orders found for the given date range.'}, status=404)

        # Prepare data for DataFrame
        data = []
        grand_total = 0
        for order in orders:
            if order.order_status != 'Cancelled':
                grand_total += float(order.total_amount)

            data.append({
                'Order #': order.reference_number,
                'Date': order.created_at.strftime('%Y-%m-%d %H:%M:%S'),
                'Customer': order.customer_name or 'Walk-in Customer',
                'Customer Phone': order.customer_phone or 'N/A',
                'Items Count': order.items.count(),
                'Status': order.order_status,
                'Payment Status': order.payment_status,
                'Payment Method': order.payment_method,
                'Subtotal': float(order.subtotal),
                'Tax': float(order.tax_amount),
                'Discount': float(order.discount_amount or 0),
                'Total': float(order.total_amount),
            })

        # Create DataFrame
        df = pd.DataFrame(data)

        # Create an Excel writer (in memory)
        buffer = BytesIO()
        with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
            # Write data
            df.to_excel(writer, index=False, sheet_name='Orders Report', startrow=4)

            # Add metadata
            workbook = writer.book
            sheet = writer.sheets['Orders Report']

            # Report title
            sheet['A1'] = 'Orders Report'
            sheet['A2'] = f"Report Period: {start_date_obj.strftime('%d %b %Y %H:%M:%S')} to {end_date_obj.strftime('%d %b %Y %H:%M:%S')}"
            if status:
                sheet['A2'].value += f" | Status: {status}"

            # Grand total
            total_row = len(df) + 6
            sheet[f'K{total_row}'] = 'GRAND TOTAL:'
            sheet[f'L{total_row}'] = grand_total

            # Basic formatting (optional)
            for cell in sheet["1:1"]:
                cell.font = cell.font.copy(bold=True)
            for col in sheet.columns:
                max_length = max(len(str(cell.value)) for cell in col)
                sheet.column_dimensions[col[0].column_letter].width = max_length + 5

        buffer.seek(0)

        # Prepare HTTP response
        filename = f"orders_report_{start_date_obj.strftime('%Y%m%d_%H%M%S')}_to_{end_date_obj.strftime('%Y%m%d_%H%M%S')}.xlsx"
        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response

    except Exception as e:
        return JsonResponse({
            'error': f'An error occurred while generating the Excel report: {str(e)}'
        }, status=500)


@login_required
@management_required
def export_order_items_excel(request):
    """Export order items as Excel file using openpyxl"""
    try:
        # Filters
        start_date = request.GET.get('start')
        end_date = request.GET.get('end')
        category = request.GET.get('category')
        today = timezone.now()

        # Date parsing
        def parse_date(date_str, is_end=False):
            if not date_str:
                return today if is_end else today - timedelta(days=30)
            try:
                dt = datetime.strptime(date_str, '%Y-%m-%d %H:%M:%S')
            except ValueError:
                dt = datetime.strptime(date_str, '%Y-%m-%d')
            if timezone.is_naive(dt):
                dt = timezone.make_aware(dt)
            if is_end:
                dt = dt.replace(hour=23, minute=59, second=59)
            return dt

        start_date_obj = parse_date(start_date)
        end_date_obj = parse_date(end_date, True)

        if start_date_obj > end_date_obj:
            return JsonResponse({'error': 'Start date cannot be after end date.'}, status=400)

        # Query order items
        product_sales = OrderItem.objects.filter(
            order__created_at__gte=start_date_obj,
            order__created_at__lte=end_date_obj,
            order__order_status='Completed'
        )

        category_name = "All Categories"
        if category:
            product_sales = product_sales.filter(product__category_id=category)
            try:
                category_name = PosCategory.objects.get(id=category).name
            except PosCategory.DoesNotExist:
                category_name = "Unknown Category"

        # Aggregation
        product_sales = product_sales.values(
            'product__name', 
            'product__category__name'
        ).annotate(
            total_quantity=Sum('quantity'),
            total_sales=Sum('total_price')
        ).order_by('-total_quantity')

        # Compute averages safely
        for p in product_sales:
            qty = p['total_quantity'] or 0
            total = p['total_sales'] or 0
            p['avg_unit_price'] = float(total) / float(qty) if qty else 0.0

        # Excel setup
        wb = Workbook()
        ws = wb.active
        ws.title = "Products Sold Report"

        # Styles
        bold = Font(bold=True)
        header_fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
        center = Alignment(horizontal="center", vertical="center")

        # Title
        ws.merge_cells('A1:E1')
        ws['A1'] = 'Products Sold Report'
        ws['A1'].font = Font(size=14, bold=True)
        ws['A1'].alignment = center

        # Date range
        ws.merge_cells('A2:E2')
        ws['A2'] = f"Report Period: {start_date_obj.strftime('%d %b %Y %H:%M:%S')} to {end_date_obj.strftime('%d %b %Y %H:%M:%S')} | Category: {category_name}"
        ws['A2'].alignment = center

        # Headers
        headers = ['Product', 'Category', 'Quantity Sold', 'Unit Price', 'Total Sales']
        ws.append(headers)
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col_num)
            cell.font = bold
            cell.fill = header_fill
            ws.column_dimensions[get_column_letter(col_num)].width = 20

        # Data rows
        grand_total_qty = 0
        grand_total_sales = 0.0
        for p in product_sales:
            ws.append([
                p['product__name'],
                p['product__category__name'] or 'Unknown',
                p['total_quantity'],
                round(p['avg_unit_price'], 2),
                round(float(p['total_sales']), 2)
            ])
            grand_total_qty += p['total_quantity']
            grand_total_sales += float(p['total_sales'] or 0.0)

        # Totals
        total_row = ws.max_row + 2
        ws.cell(row=total_row, column=2, value="GRAND TOTAL:").font = bold
        ws.cell(row=total_row, column=3, value=grand_total_qty).font = bold
        ws.cell(row=total_row, column=5, value=grand_total_sales).font = bold

        # Save to memory
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        # Response
        filename = f"products_sold_report_{start_date_obj.strftime('%Y%m%d_%H%M%S')}_to_{end_date_obj.strftime('%Y%m%d_%H%M%S')}.xlsx"
        response = HttpResponse(output.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response

    except Exception as e:
        return JsonResponse({'error': f'Error generating Excel: {str(e)}'}, status=500)


@login_required
def sales_receipt(request):
    """Display a printable receipt for sales summary report"""
    if not can_access_management(request.user):
        messages.error(request, 'You do not have permission to access this page.')
        return redirect('dashboard')
    
    try:
        # Get date range from request
        start_date_param = request.GET.get('start_date')
        end_date_param = request.GET.get('end_date')
        user_id_param = request.GET.get('user_id')  # Get user_id parameter
        redirect_to = request.GET.get('redirect_to', '')  # Get redirect_to parameter
        
        # Set default date range to current month if not provided
        today = timezone.now()
        
        if not start_date_param or not end_date_param:
            start_date = timezone.make_aware(datetime.combine(today.replace(day=1).date(), datetime.min.time()))
            end_date = timezone.make_aware(datetime.combine(today.date(), datetime.max.time()))
        else:
            try:
                # Try to parse as datetime with time first
                try:
                    start_date = datetime.strptime(start_date_param, '%Y-%m-%d %H:%M:%S')
                    # Make timezone aware
                    if timezone.is_naive(start_date):
                        start_date = timezone.make_aware(start_date)
                except ValueError:
                    # Fall back to date only format
                    start_date = datetime.strptime(start_date_param, '%Y-%m-%d').date()
                    # Convert to datetime at start of day
                    start_date = timezone.make_aware(datetime.combine(start_date, datetime.min.time()))
                    
                # Try to parse as datetime with time first
                try:
                    end_date = datetime.strptime(end_date_param, '%Y-%m-%d %H:%M:%S')
                    # Make timezone aware
                    if timezone.is_naive(end_date):
                        end_date = timezone.make_aware(end_date)
                except ValueError:
                    # Fall back to date only format
                    end_date = datetime.strptime(end_date_param, '%Y-%m-%d').date()
                    # Convert to datetime at end of day
                    end_date = timezone.make_aware(datetime.combine(end_date, datetime.max.time()))
            except ValueError:
                # If there's any issue parsing the dates, use default values
                start_date = timezone.make_aware(datetime.combine(today.replace(day=1).date(), datetime.min.time()))
                end_date = timezone.make_aware(datetime.combine(today.date(), datetime.max.time()))
        
        # Create a cache key based on the date range and user filter
        cache_key = f"sales_receipt_{start_date.strftime('%Y%m%d_%H%M%S')}_{end_date.strftime('%Y%m%d_%H%M%S')}"
        if user_id_param:
            cache_key += f"_user{user_id_param}"
        
        # Try to get cached results (only if no user filter, since user-specific reports shouldn't be cached)
        cached_data = None
        if not user_id_param:
            cached_data = cache.get(cache_key)
            if cached_data:
                logger.debug(f"Returning cached sales receipt data for {start_date} to {end_date}")
                # Add redirect_to to cached data
                cached_data['redirect_to'] = redirect_to
                return render(request, 'posapp/reports/sales_receipt.html', cached_data)
        
        # SPECIAL HANDLING: If this is called from end_day redirect and no orders in range,
        # it might be because the EndDay was just created. Let's check for the correct range.
        orders_in_range = Order.objects.filter(
            created_at__gte=start_date,
            created_at__lte=end_date
        ).count()
        
        if orders_in_range == 0 and redirect_to == 'dashboard':
            # This might be from end_day redirect with wrong date range
            # Let's try to get the correct business day range
            
            # Get the last two end days
            end_days = list(EndDay.objects.order_by('-end_date')[:2])
            if len(end_days) >= 2:
                # Use the range from second-to-last to last end day
                corrected_start_date = end_days[1].end_date  # Second-to-last
                corrected_end_date = end_days[0].end_date    # Last (just created)
                
                corrected_orders = Order.objects.filter(
                    created_at__gte=corrected_start_date,
                    created_at__lte=corrected_end_date
                ).count()
                
                if corrected_orders > 0:
                    # Use the corrected date range
                    start_date = corrected_start_date
                    end_date = corrected_end_date
            elif len(end_days) == 1:
                # Only one end day exists, use from beginning of time to this end day
                corrected_start_date = timezone.make_aware(datetime(2020, 1, 1))  # Very old date
                corrected_end_date = end_days[0].end_date
                
                corrected_orders = Order.objects.filter(
                    created_at__gte=corrected_start_date,
                    created_at__lte=corrected_end_date
                ).count()
                
                if corrected_orders > 0:
                    # Use the corrected date range
                    start_date = corrected_start_date
                    end_date = corrected_end_date
        
        # Get all completed orders in the date range (potentially corrected)
        completed_orders = Order.objects.filter(
            created_at__gte=start_date,
            created_at__lte=end_date,
            order_status='Completed'
        )
        

        
        # Filter by user if user_id is provided
        if user_id_param:
            try:
                user_id = int(user_id_param)
                completed_orders = completed_orders.filter(user_id=user_id)
            except (ValueError, TypeError):
                # Invalid user_id, ignore the filter
                pass
        
        # Calculate revenue breakdown components
        revenue_breakdown = completed_orders.aggregate(
            total_amount=Coalesce(Sum('total_amount'), Decimal('0.00')),
            sales_revenue=Coalesce(Sum('subtotal'), Decimal('0.00')),  # Base sales without service charge
            service_charge_total=Coalesce(Sum('service_charge_amount'), Decimal('0.00'))  # Service charges
        )
        
        # Extract individual components
        total_sales = revenue_breakdown['total_amount']
        sales_revenue = revenue_breakdown['sales_revenue']  # This is the base sales amount
        total_service_charge = revenue_breakdown['service_charge_total']
        
        # Calculate the total paid amount
        total_paid = completed_orders.filter(
            payment_status='Paid'
        ).aggregate(
            total=Coalesce(Sum('total_amount'), Decimal('0.00'))
        )['total']
        
        # Calculate the total pending amount
        total_pending = completed_orders.filter(
            payment_status='Pending'
        ).aggregate(
            total=Coalesce(Sum('total_amount'), Decimal('0.00'))
        )['total']
        
        # Calculate order type breakdown
        takeaway_orders = completed_orders.filter(order_type='Takeaway')
        takeaway_stats = takeaway_orders.aggregate(
            count=Count('id'),
            total_amount=Coalesce(Sum('total_amount'), Decimal('0.00'))
        )
        
        dinein_orders = completed_orders.filter(order_type='Dine In')
        dinein_stats = dinein_orders.aggregate(
            count=Count('id'),
            total_amount=Coalesce(Sum('total_amount'), Decimal('0.00'))
        )
        
        delivery_orders = completed_orders.filter(order_type='Delivery')
        delivery_stats = delivery_orders.aggregate(
            count=Count('id'),
            total_amount=Coalesce(Sum('total_amount'), Decimal('0.00'))
        )
        
        # Calculate delivery person breakdown
        delivery_person_stats = delivery_orders.values('delivery_person__name').annotate(
            order_count=Count('id'),
            total_amount=Sum('total_amount')
        ).order_by('-total_amount')
        
        # Handle null delivery person (orders without assigned delivery person)
        delivery_person_stats = list(delivery_person_stats)
        for stat in delivery_person_stats:
            if stat['delivery_person__name'] is None:
                stat['delivery_person__name'] = 'Unassigned'
        
        # Get all bill adjustments in the date range
        bill_adjustments = BillAdjustment.objects.filter(
            created_at__gte=start_date,
            created_at__lte=end_date
        )
        
        # Calculate the total bill adjustments
        total_bill_adjustments = bill_adjustments.aggregate(
            total=Coalesce(Sum('price'), Decimal('0.00'))
        )['total']
        
        # Get all advance adjustments in the date range
        advance_adjustments = AdvanceAdjustment.objects.filter(
            created_at__gte=start_date,
            created_at__lte=end_date
        )
        
        # Calculate the total advance adjustments
        total_advance_adjustments = advance_adjustments.aggregate(
            total=Coalesce(Sum('amount'), Decimal('0.00'))
        )['total']
        
        # Calculate the total adjustments
        total_adjustments = total_bill_adjustments + total_advance_adjustments
        
        # Calculate total discounts given in completed orders
        total_discounts = completed_orders.aggregate(
            total=Coalesce(Sum('discount_amount'), Decimal('0.00'))
        )['total']
        
        # Calculate total delivery charges from completed orders
        total_delivery_charges = completed_orders.aggregate(
            total=Coalesce(Sum('delivery_charges'), Decimal('0.00'))
        )['total']
        
        # Calculate the net revenue: total_sales - service_charge - discounts - adjustments - delivery_charges
        # Starting with completed sales (total_sales) which includes all charges
        # Subtracting service charges, discounts, adjustments, and delivery charges
        net_revenue = total_sales - total_service_charge - total_discounts - total_adjustments - total_delivery_charges
        
        # Calculate total deductions for detailed breakdown
        total_deductions = total_service_charge + total_discounts + total_adjustments + total_delivery_charges
        
        # Determine if there's a shortage
        is_shortage = net_revenue < 0
        shortage_amount = abs(net_revenue) if is_shortage else Decimal('0.00')
        
        # Get products sold in the date range
        products_sold_query = OrderItem.objects.filter(
            order__created_at__gte=start_date,
            order__created_at__lte=end_date,
            order__order_status='Completed'
        )
        
        # Filter by user if user_id is provided
        if user_id_param:
            try:
                user_id = int(user_id_param)
                products_sold_query = products_sold_query.filter(order__user_id=user_id)
            except (ValueError, TypeError):
                # Invalid user_id, ignore the filter
                pass
        
        products_sold = products_sold_query.values(
            'product__name'
        ).annotate(
            total_quantity=Sum('quantity'),
            total_sales=Sum('total_price')
        ).order_by('-total_quantity')
        
        # Get user information if user_id is provided
        filtered_user = None
        if user_id_param:
            try:
                from django.contrib.auth.models import User
                user_id = int(user_id_param)
                filtered_user = User.objects.get(id=user_id)
            except (ValueError, TypeError, User.DoesNotExist):
                # Invalid user_id or user not found, ignore
                pass
        
        # Prepare context data
        context = {
            'completed_orders': completed_orders,
            'total_sales': total_sales,
            'sales_revenue': sales_revenue,  # Add sales revenue for breakdown
            'total_service_charge': total_service_charge,
            'total_delivery_charges': total_delivery_charges,  # Add total delivery charges
            'total_paid': total_paid,
            'total_pending': total_pending,
            'bill_adjustments': bill_adjustments,
            'total_bill_adjustments': total_bill_adjustments,
            'advance_adjustments': advance_adjustments,
            'total_advance_adjustments': total_advance_adjustments,
            'total_adjustments': total_adjustments,
            'total_discounts': total_discounts,  # Add total discounts
            'total_deductions': total_deductions,  # Add total deductions
            'net_revenue': net_revenue,
            'is_shortage': is_shortage,
            'shortage_amount': shortage_amount,
            'products_sold': products_sold,
            'start_date': start_date,
            'end_date': end_date,
            'now': timezone.now(),
            'redirect_to': redirect_to,  # Add redirect_to to context
            'filtered_user': filtered_user,  # Add user info for filtered reports
            # Order type breakdown
            'takeaway_stats': takeaway_stats,
            'dinein_stats': dinein_stats,
            'delivery_stats': delivery_stats,
            'delivery_person_stats': delivery_person_stats,
        }
        
        # Get business settings
        try:
            from ..views.settings_views import get_or_create_settings as get_settings
            business_settings = get_settings([
                'business_name', 'business_address', 'business_phone', 
                'business_email', 'currency_symbol'
            ])
            
            context.update({
                'business_name': business_settings['business_name'].setting_value,
                'business_address': business_settings['business_address'].setting_value,
                'business_phone': business_settings['business_phone'].setting_value,
                'business_email': business_settings['business_email'].setting_value,
                'currency_symbol': business_settings['currency_symbol'].setting_value or '$',
            })
            
            # Get business logo from BusinessLogo model
            context['business_logo'] = BusinessLogo.get_logo_url()
        except Exception as e:
            logger.error(f"Error getting business settings: {str(e)}")
            context.update({
                'business_name': 'POS System',
                'currency_symbol': '$',
            })
        
        # Get receipt settings
        try:
            receipt_settings = get_settings([
                'receipt_header', 'receipt_footer', 'receipt_show_logo',
                'receipt_show_cashier', 'receipt_paper_size',
                'receipt_custom_css'
            ])
            
            context.update({
                'receipt_header': receipt_settings['receipt_header'].setting_value,
                'receipt_footer': receipt_settings['receipt_footer'].setting_value,
                'receipt_show_logo': receipt_settings['receipt_show_logo'].setting_value == 'True',
                'receipt_show_cashier': receipt_settings['receipt_show_cashier'].setting_value == 'True',
                'receipt_paper_size': receipt_settings['receipt_paper_size'].setting_value,
                'receipt_custom_css': receipt_settings['receipt_custom_css'].setting_value,
            })
        except Exception as e:
            logger.error(f"Error getting receipt settings: {str(e)}")
        
        # Store in cache for 1 hour (3600 seconds), but don't include redirect_to or filtered_user in cache
        # Also don't cache user-specific reports
        if not user_id_param:
            cache_context = context.copy()
            if 'redirect_to' in cache_context:
                del cache_context['redirect_to']
            if 'filtered_user' in cache_context:
                del cache_context['filtered_user']
            cache.set(cache_key, cache_context, 3600)
        
        logger.info(f"Generated sales receipt for period {start_date} to {end_date} by user {request.user.username}")
        
        return render(request, 'posapp/reports/sales_receipt.html', context)
    except Exception as e:
        logger.exception(f"Error generating sales receipt: {str(e)}")
        messages.error(request, f"An error occurred while generating the sales receipt: {str(e)}")
        return redirect('reports_dashboard')





@login_required
@management_required
def end_day_reports(request):
    """Display list of all end day reports"""
    user_is_admin = is_admin(request.user)
    
    # Get all end days ordered by oldest first
    end_days = list(EndDay.objects.all().order_by('end_date'))
    
    # Date filter (admin only)
    start_date = request.GET.get('start_date', '')
    end_date = request.GET.get('end_date', '')
    if user_is_admin:
        if start_date:
            try:
                start_date_dt = timezone.datetime.strptime(start_date, '%Y-%m-%d').date()
                end_days = [ed for ed in end_days if ed.end_date.date() >= start_date_dt]
            except ValueError:
                pass
        if end_date:
            try:
                end_date_dt = timezone.datetime.strptime(end_date, '%Y-%m-%d').date()
                end_days = [ed for ed in end_days if ed.end_date.date() <= end_date_dt]
            except ValueError:
                pass
    
    # Get earliest order date for the very first period
    earliest_order = Order.objects.order_by('created_at').first()
    earliest_order_date = earliest_order.created_at if earliest_order else None
    
    end_days_with_periods = []
    for i, end_day in enumerate(end_days):
        if i == 0:
            period_start = earliest_order_date
        else:
            period_start = end_days[i-1].end_date
        period_end = end_day.end_date

        orders = Order.objects.filter(
            created_at__gte=period_start,
            created_at__lte=period_end
        ).exclude(order_status='Cancelled')

        # Get adjustments for this period
        bill_adjustments = BillAdjustment.objects.filter(
            created_at__gte=period_start,
            created_at__lte=period_end
        )
        advance_adjustments = AdvanceAdjustment.objects.filter(
            created_at__gte=period_start,
            created_at__lte=period_end
        )

        bill_adjustment_total = bill_adjustments.aggregate(Sum('price'))['price__sum'] or 0
        advance_adjustment_total = advance_adjustments.aggregate(Sum('amount'))['amount__sum'] or 0
        adjustment_total = bill_adjustment_total + advance_adjustment_total

        order_total = orders.aggregate(Sum('total_amount'))['total_amount__sum'] or 0
        net_total = order_total - adjustment_total

        end_days_with_periods.append({
            'end_day': end_day,
            'period_start': period_start,
            'period_end': period_end,
            'orders_count': orders.count(),
            'order_total': order_total,
            'bill_adjustment_total': bill_adjustment_total,
            'advance_adjustment_total': advance_adjustment_total,
            'adjustment_total': adjustment_total,
            'net_total': net_total
        })
    
    # Show newest first in the template
    end_days_with_periods.reverse()
    
    context = {
        'end_days_data': end_days_with_periods,
        'is_admin': user_is_admin,
    }
    return render(request, 'posapp/reports/end_day_reports.html', context) 